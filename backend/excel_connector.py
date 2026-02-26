    #import xlwings as xw
import openpyxl
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
import pandas as pd
import yfinance as yf
import datetime
import threading
from zoneinfo import ZoneInfo
import logging
from dotenv import load_dotenv 
import os

from logger_setup import timezone

logger = logging.getLogger('backend')


class Portfolio:
    def __init__(self):
        
        load_dotenv()
        self.fullname = os.getenv("PATH_NAME")

        self.portfolio_df = pd.DataFrame()
        self.stocks_dict = {}
        self.wb = None
        self.money_invested_sum = 0
        self.portfolio_start_date = None
        self.portfolio_value:float = 0.00
        self.money_invested = []
        self.stock_tickers = []
        self.stock_names = []
        self.quantity_list = []
        self.start_dates = []
        self.sector = []
        self.risk = []

        #LOG
        self.log_df = pd.DataFrame()

        #BETA SHEET
        self.beta_dict = {}

        #LEDGER
        self.investor_df_og = pd.DataFrame()
        self.money_added_list = []

        #RISK TABLE
        self.risk_table_total_investment = 0

        #COMPONENTS FOR DASH
        self.allocation_funds_portfolio_df  = pd.DataFrame()
        self.sector_funds_portfolio_df  = pd.DataFrame()
        self.funds_sector_returns_df = pd.DataFrame()
        self.combined_allocation_df = pd.DataFrame()
        self.combined_sector_df = pd.DataFrame()
        self.allocation_sector_df = pd.DataFrame()
        self.combined_sector_returns_df = pd.DataFrame()
        self.risk_table_df = pd.DataFrame()

        self.is_updating = False
        self.file_lock = threading.Lock()
    
    def safe_save(self, wb):
        with self.file_lock:
            try:
                wb.save(self.fullname)
                logger.debug(f"Successfully saved: {wb}")
                return "Successfully saved the workbook!"
            except PermissionError:
                logger.debug("Error", exc_info=True)
                return "Could not save. Please close the Excel file if it's open \
                    or ensure that the location in the .env file is accurate"
            except Exception as e:
                logger.debug("Error", exc_info=True)
                return f"An unexpected error occurred while saving: {e}"

        
    def final_safe_save(self, wb=None):
        """Ensures only one thread writes to the file at a time"""
        if self.is_updating is False:
            if wb is None:
                wb = self.safe_load()
            with self.file_lock:
                try:
                    wb.save(self.fullname)
                    logger.debug(f"Successfully saved: {wb}")
                    return "Successfully saved the workbook!"
                except PermissionError:
                    logger.debug("Error", exc_info=True)
                    return "Could not save. Please close the Excel file if it's open \
                        or ensure that the location in the .env file is accurate"
                except Exception as e:
                    logger.debug("Error", exc_info=True)
                    return f"An unexpected error occurred while saving: {e}"
        else:
            return "Cannot save the workbook at the moment. Please wait a few seconds and try again."
            logger.debug("Didn't save workbook")


    def safe_load(self, data_only=False):
        """Ensures file isn't read while another thread is writing"""
        with self.file_lock:
            return openpyxl.load_workbook(self.fullname, data_only=data_only)


    def row_iterator_and_writer (self, sheet, start_cell, values:list):
        # Start_Row 3 is the 3rd row
        col_str, row = coordinate_from_string(start_cell)
        col = column_index_from_string(col_str)

        for i, value in enumerate(values):
            sheet.cell(row=row + i, column=col).value = value


    def add_data_to_portfolio(self, wb=None):
        try:
            if wb is None:
                wb = self.safe_load(data_only=False)

            sh_port = wb['Portfolio']
            sh_funds = wb['Funds_Portfolio']
            sh_ledger = wb['Ledger']

            # Initialize lists (11 empty lists)
            (current_price, change, two_hundred_day_average, market_cap, fifty_two_week_high, 
            fifty_two_week_low, investment_value, today_profit_loss, today_profit_loss_perc, 
            total_profit_loss, total_profit_loss_perc) = [[] for _ in range(11)]

            portfolio_sum = float(sh_funds['A7'].value)

            logger.info("Retrieving information from Yahoo Finance...")

            # Data collection loop
            for i, name in enumerate(self.stock_names):
                temp_stock = yf.Ticker(name)
                info = temp_stock.get_info()
                
                curr = info.get('regularMarketPrice', 0)
                chg = info.get('regularMarketChange', 0)
                
                current_price.append(curr)
                change.append(chg)
                two_hundred_day_average.append(info.get('twoHundredDayAverage', 0))
                
                # Calculations
                inv_val = curr * self.quantity_list[i]
                investment_value.append(inv_val)
                
                today_pl = chg * self.quantity_list[i]
                today_profit_loss.append(today_pl)
                
                # today_profit_loss_perc logic: curr / (curr - chg) - 1
                denom = (curr - chg)
                today_profit_loss_perc.append(curr / denom - 1 if denom != 0 else 0)
                
                # total_profit_loss logic: inv_val - money_invested[i]
                t_pl = inv_val - self.money_invested[i]
                total_profit_loss.append(t_pl)
                
                # total_profit_loss_perc logic: t_pl / money_invested[i]
                total_profit_loss_perc.append(t_pl / self.money_invested[i] if self.money_invested[i] != 0 else 0)
                
                fifty_two_week_high.append(info.get('fiftyTwoWeekHigh', 0))
                fifty_two_week_low.append(info.get('fiftyTwoWeekLow', 0))
        
        except Exception as e:
            print("Caught an error.")
            logger.debug(f"Error {e}", exc_info=True)

        try:
            logger.debug("Writing each column of Portfolio sheet")
            #sheet1['C3'].options(transpose=True).value = stock_names
            self.row_iterator_and_writer(sh_port, 'C3', values = self.stock_names)

            #sheet1['D3'].options(transpose=True).value = start_dates
            self.row_iterator_and_writer(sh_port, 'D3', values=self.start_dates)

            #sheet1['E3'].options(transpose=True).value = change
            self.row_iterator_and_writer(sh_port, "E3", values=change)

            #sheet1['F3'].options(transpose=True).value = current_price
            self.row_iterator_and_writer(sh_port, 'F3', values=current_price)

            #sheet1['G3'].options(transpose=True).value = quantity_list
            self.row_iterator_and_writer(sh_port, 'G3', values=self.quantity_list)

            #sheet1['H3'].options(transpose=True).value = two_hundred_day_average
            self.row_iterator_and_writer(sh_port, 'H3', values=two_hundred_day_average)

            #sheet1['I3'].options(transpose=True).value = money_invested
            self.row_iterator_and_writer(sh_port, 'I3', values=self.money_invested)

            #sheet1['J3'].options(transpose=True).value = investment_value
            self.row_iterator_and_writer(sh_port, 'J3', values=investment_value)

            #sheet1['L3'].options(transpose=True).value = today_profit_loss
            self.row_iterator_and_writer(sh_port, 'L3', values=today_profit_loss)

            #sheet1['M3'].options(transpose=True).value = today_profit_loss_perc
            self.row_iterator_and_writer(sh_port, 'M3', values=today_profit_loss_perc)

            #sheet1['N3'].options(transpose=True).value = total_profit_loss
            self.row_iterator_and_writer(sh_port, 'N3', values=today_profit_loss)

            #sheet1['O3'].options(transpose=True).value = total_profit_loss_perc
            self.row_iterator_and_writer(sh_port, 'O3', values=total_profit_loss_perc)

            #sheet1['P3'].options(transpose=True).value = fifty_two_week_high
            self.row_iterator_and_writer(sh_port, 'P3', values=fifty_two_week_high)

            #sheet1['Q3'].options(transpose=True).value = fifty_two_week_low
            self.row_iterator_and_writer(sh_port, 'Q3', values=fifty_two_week_low)

            #sheet1['R3'].options(transpose=True).value = sector
            self.row_iterator_and_writer(sh_port, 'R3', values=self.sector)

            #sheet1['S3'].options(transpose=True).value = risk
            self.row_iterator_and_writer(sh_port, 'S3', values=self.risk)

            # Date and Time Update
            date_and_time = datetime.datetime.now(ZoneInfo(timezone))
            sh_port['A2'] = date_and_time.strftime('%d/%m/%Y %I:%M %p')

            # Portfolio Sum Calculations 
            portfolio_sum += sum(investment_value)
            sh_port['A4'] = portfolio_sum

            # Money Invested Sum 
            money_invested_sum = float(sh_funds['A4'].value or 0)
            money_invested_sum += sum(self.money_invested)
            sh_port['A7'] = money_invested_sum

            # Total Stocks 
            total_stocks = float(sh_funds['A13'].value or 0)
            total_stocks += sum(self.quantity_list)
            sh_port['A9'] = "Total Stocks"
            sh_port['A10'] = total_stocks

            # Returns 
            total_profit = float(sh_funds['A10'].value or 0)
            total_profit += sum(total_profit_loss)
            sh_port['A14'] = total_profit

        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)
            print("Caught an error.")

        try:
            # Number of Months 
            last_row_ledger = sh_ledger.max_row
            while last_row_ledger > 1 and sh_ledger.cell(row=last_row_ledger, column=1).value is None:
                last_row_ledger -= 1
            sh_port['A19'] = last_row_ledger - 4

            # Allocation Logic
            last_row_portfolio = sh_port.max_row
            while last_row_portfolio > 1 and sh_port.cell(row=last_row_portfolio, column=3).value is None:
                last_row_portfolio -= 1
                
            for r in range(3, last_row_portfolio + 1):
                cell = sh_port.cell(row=r, column=11) # Column K
                cell.value = f'=J{r}/A$4'
                cell.number_format = "0.00%"

            # NAV Logic
            sh_port['A24'] = "NAV"
            sh_port['A25'] = '=A4/A10'

            self.portfolio_df['Investment Value'] = investment_value

            self.safe_save(wb)
        
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)
            print("Caught an error")

    
    
    def df_for_dash(self, wb=None):
        """Creating DFs to display on Dash"""
        try:

            if wb is None:
                wb = self.safe_load(data_only=True)
            
            allocation_df = self.portfolio_df[['Symbol', 'Allocation']]
            self.combined_allocation_df = pd.concat([allocation_df, self.allocation_funds_portfolio_df], ignore_index=True)
            self.combined_allocation_df['Allocation'] = self.combined_allocation_df['Allocation'].astype(float)
            self.combined_allocation_df.loc[:,'Allocation'] = self.combined_allocation_df.loc[:,'Allocation'] * 100

            sector_df = self.portfolio_df[['Symbol', 'Sector', 'Allocation']]
            self.combined_sector_df = pd.concat([sector_df, self.sector_funds_portfolio_df], ignore_index=True)
            self.combined_sector_df['Allocation'] = self.combined_sector_df['Allocation'].astype(float)
            self.combined_sector_df.loc[:,'Allocation'] = self.combined_sector_df.loc[:,'Allocation'] * 100

            sector_returns_df = self.portfolio_df[['Symbol', 'Sector', 'Today Profit and Loss (Percentage)', 'Total Profit and Loss (Percentage)']]
            self.combined_sector_returns_df = pd.concat([sector_returns_df, self.funds_sector_returns_df], ignore_index=True)
            self.combined_sector_returns_df.loc[:,'Today Profit and Loss (Percentage)'] = 100 * self.combined_sector_returns_df.loc[:,'Today Profit and Loss (Percentage)']
            self.combined_sector_returns_df.loc[:,'Total Profit and Loss (Percentage)'] = 100 * self.combined_sector_returns_df.loc[:,'Total Profit and Loss (Percentage)']
        
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)
            print("Caught an error.")
        
        try:
            ##RISK TABLE ##
            sh_ledger = wb['Ledger']
            sh_funds = wb['Funds_Portfolio']

            low_risk_stocks = self.portfolio_df[self.portfolio_df['Risk'].str.upper() == "LOW"]
            high_risk_stocks = self.portfolio_df[self.portfolio_df['Risk'].str.upper() == "HIGH"]
            
            low_risk_stocks_sum = float(sh_funds['A16'].value or 0)
            high_risk_stocks_sum = float(sh_funds['A19'].value or 0)

            low_risk_stocks_sum += low_risk_stocks['Investment Value'].sum()
            high_risk_stocks_sum += high_risk_stocks['Investment Value'].sum()

            logger.debug(low_risk_stocks_sum)

            self.portfolio_value = float(wb['Portfolio']['A4'].value)

            self.risk_table_df.loc[0, 'Current Value'] = low_risk_stocks_sum
            self.risk_table_df.loc[1, 'Current Value'] = high_risk_stocks_sum

            self.risk_table_df.loc[0, 'Current %'] = low_risk_stocks_sum/self.portfolio_value
            self.risk_table_df.loc[1, 'Current %'] = high_risk_stocks_sum/self.portfolio_value

            self.risk_table_df.loc[0, 'Allocation Value'] = 0.5 * self.risk_table_total_investment 
            self.risk_table_df.loc[1, 'Allocation Value'] = 0.5 * self.risk_table_total_investment 
            
            self.risk_table_df.loc[0, 'To reduce/add']  = 0.5 * self.risk_table_total_investment - low_risk_stocks_sum 
            self.risk_table_df.loc[1, 'To reduce/add']  = 0.5 * self.risk_table_total_investment - high_risk_stocks_sum
        
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)
            print("Caught an error.")
        


    def excel_reader(self, wb=None):
        try:

            logger.debug("Starting excel reader")
            
            if wb is None:
                wb = self.safe_load(data_only=True) # data_only=True gets formula results
            
            sheet = wb['Portfolio']

            self.portfolio_value = float(wb['Portfolio']['A4'].value)

            # Identify the last used row in Column C
            last_row = sheet.max_row
            while last_row > 1 and sheet.cell(row=last_row, column=3).value is None:
                last_row -= 1

            # Extract Range C2:S{last_row} into a list of lists
            # Column C is 3, Column S is 19
            data = []
            for row in sheet.iter_rows(min_row=2, max_row=last_row, min_col=3, max_col=19, values_only=True):
                data.append(row)

            # Convert to DataFrame (using the first row as header)
            raw_df = pd.DataFrame(data)
            header_cols = raw_df.iloc[0]

            self.portfolio_df = raw_df[1:].copy()
            self.portfolio_df.columns = header_cols
            self.portfolio_df = self.portfolio_df.reset_index(drop=True)

            cell = sheet['K3']
            
            #Test to ensure values have been read from Excel & not formulae
            if cell.value is None:
                return 1
            
            self.money_invested_sum = float(sheet['A7'].value or 0)
            self.portfolio_start_date = sheet['A17'].value

            # Build the stocks dictionary
            self.stocks_dict = {}

            for _, row in self.portfolio_df.iterrows():
                self.stocks_dict[row['Symbol']] = [
                    float(row['Amount']),
                    float(row['Money Invested'])
                ]
            
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)

        try:
            # Prepare return lists
            self.stock_names = self.portfolio_df['Symbol'].tolist()
            self.quantity_list = self.portfolio_df['Amount'].tolist()
            self.money_invested = self.portfolio_df['Money Invested'].tolist()
            self.start_dates = self.portfolio_df['Start Date'].tolist()
            self.sector = self.portfolio_df['Sector'].tolist()
            self.risk = self.portfolio_df['Risk'].tolist()
            self.stock_tickers = ', '.join(map(str, self.stock_names))

            sh_funds = wb['Funds_Portfolio']
            last_row_funds = sh_funds.max_row
            while last_row_funds > 1 and sh_funds.cell(row=last_row_funds, column=3).value is None:
                last_row_funds -= 1

            funds_data = []
            for row in sh_funds.iter_rows(min_row=3, max_row=last_row_funds, min_col=3, max_col=19, values_only=True):
                funds_data.append(row)
            
            funds_raw_df = pd.DataFrame(funds_data)
            
            header_cols = funds_raw_df.iloc[0]
            funds_portfolio_df = funds_raw_df[1:].copy()
            funds_portfolio_df.columns = header_cols
            funds_portfolio_df = funds_portfolio_df.reset_index(drop=True)

            self.allocation_funds_portfolio_df = funds_portfolio_df[['Symbol', 'Allocation']]

            allocation_df = self.portfolio_df[['Symbol', 'Allocation']]

            self.combined_allocation_df = pd.concat([allocation_df, self.allocation_funds_portfolio_df], ignore_index=True)
            
            self.combined_allocation_df['Allocation'] = self.combined_allocation_df['Allocation'].astype(float)
            
            self.combined_allocation_df.loc[:,'Allocation'] = self.combined_allocation_df.loc[:,'Allocation'] * 100

        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)

        try:
            if self.combined_allocation_df['Allocation'].dtype != float:
                logger.info("The app didn't read the Excel tables accurately.\
                    Please manually save the Excel file before running this app.")
        except:
            logger.info("Error", exc_info=True)

        try:
            #CREATING DF CONTAINING STOCK'S SECTOR'S SHARE OF PORTFOLIO 
            sector_df = self.portfolio_df[['Symbol', 'Sector', 'Allocation']]
            
            self.sector_funds_portfolio_df = funds_portfolio_df[['Symbol', 'Sector', 'Allocation']]
            
            self.combined_sector_df = pd.concat([sector_df, self.sector_funds_portfolio_df], ignore_index=True)
            self.combined_sector_df.loc[:,'Allocation'] = self.combined_sector_df.loc[:,'Allocation'] * 100
            
            #CREATING DF CONTAINING STOCK'S RETURNS
            sector_returns_df = self.portfolio_df[['Symbol', 'Sector', 'Today Profit and Loss (Percentage)', 'Total Profit and Loss (Percentage)']]
            self.funds_sector_returns_df = funds_portfolio_df[['Symbol', 'Sector', 'Today Profit and Loss (Percentage)', 'Total Profit and Loss (Percentage)']]
                    
            self.funds_sector_returns_df['Today Profit and Loss (Percentage)'] = pd.to_numeric(self.funds_sector_returns_df['Today Profit and Loss (Percentage)'], errors='coerce')
            self.funds_sector_returns_df['Today Profit and Loss (Percentage)'] = self.funds_sector_returns_df['Today Profit and Loss (Percentage)'].fillna(0).astype(float)
            self.combined_sector_returns_df = pd.concat([sector_returns_df, self.funds_sector_returns_df], ignore_index=True)
            self.combined_sector_returns_df.loc[:,'Today Profit and Loss (Percentage)'] = 100 * self.combined_sector_returns_df.loc[:,'Today Profit and Loss (Percentage)']
            self.combined_sector_returns_df.loc[:,'Total Profit and Loss (Percentage)'] = 100 * self.combined_sector_returns_df.loc[:,'Total Profit and Loss (Percentage)']


            #CREATING DATAFRAME CONTAINING INVESTOR INFORMATION
            ledger_sheet = wb['Ledger']

            investor_data = []
            for row in ledger_sheet.iter_rows(min_row=4, max_row=9, min_col=9, max_col=16, values_only=True):
                investor_data.append(row)
            
            investor_raw_df = pd.DataFrame(investor_data)
            
            header_cols = investor_raw_df.iloc[0]
            self.investor_df_og = investor_raw_df[1:].copy()
            self.investor_df_og.columns = header_cols
            self.investor_df_og = self.investor_df_og.reset_index(drop=True)

            self.investor_df_og['Amount Invested'] = pd.to_numeric(self.investor_df_og['Amount Invested'], errors="coerce")

            # Setup internal variables
            self.ledger_dict = dict(zip(self.investor_df_og['Investor'], self.investor_df_og['Amount Invested'].tolist()))
            self.money_added_list = self.investor_df_og['Amount Invested'].tolist()

            risk_data = []
            for row in ledger_sheet.iter_rows(min_row=6, max_row=8, min_col=18, max_col=23, values_only=True):
                risk_data.append(row)
                    
            risk_table_raw_df = pd.DataFrame(risk_data)
                    
            header_cols = risk_table_raw_df.iloc[0]
            self.risk_table_df = risk_table_raw_df[1:].copy()
            self.risk_table_df.columns = header_cols
            self.risk_table_df = self.risk_table_df.reset_index(drop=True)
        
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)

        try:
            ## BETA SHEET ##
            for name in self.stock_names:
                sh = wb[name]

                last_row = sh.max_row
                while last_row > 1 and sh.cell(row=last_row, column=8).value is None:
                    last_row -= 1
                    
                stock_data = []
                for row in sh.iter_rows(min_row=4, max_row=last_row, min_col=8, max_col=10, values_only=True):
                    stock_data.append(row)
                
                self.beta_dict[name] = stock_data
        
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)
        
        ## LOG
        try:
            last_row_log = wb['Log'].max_row
            while last_row_log > 1 and wb['Log'].cell(row=last_row_log, column=4).value is None:
                last_row_log -= 1

            sheet = wb['Log']

            log_data = []
            for row in sheet.iter_rows(min_row=4, max_row=last_row_log, min_col=3, max_col=9, values_only=True):
                log_data.append(row)
            
            log_raw_df = pd.DataFrame(log_data)
            
            header_cols = log_raw_df.iloc[0]
            self.log_df = log_raw_df[1:].copy()
            self.log_df.columns = header_cols
            self.log_df = self.log_df.reset_index(drop=True)

            return 0

        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)


    def retrieve_last_update(self, wb=None):
        try:
            if wb is None:
                wb = self.safe_load(data_only=True)
            
            logger.info("Successfully loaded workbook")
            
            date_and_time = wb['Portfolio']['A2'].value
            
            portfolio_value = float(wb['Portfolio']['A4'].value)
            portfolio_value = f"₹ {portfolio_value:,.2f}"

            invested_amount = float(wb['Portfolio']['A7'].value)
            invested_amount = f"₹ {invested_amount:,.2f}"

            portfolio_return = float(wb['Portfolio']['A14'].value)
            portfolio_return = f"₹ {portfolio_return:,.2f}"

            portfolio_return_perc = float(wb['Portfolio']['A4'].value/wb['Portfolio']['A7'].value - 1)
            portfolio_return_perc = f"{portfolio_return_perc:.2%}"

            return date_and_time, portfolio_value, invested_amount, portfolio_return, portfolio_return_perc
        
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)



    def risk_table(self, wb=None):
        try:
            logger.debug("Starting risk_table")

            if wb is None:
                wb = self.safe_load()
            
            sh_portfolio = wb['Portfolio']
            sh_ledger = wb['Ledger']
            sh_funds = wb['Funds_Portfolio']

            # Basic Cell Updates 
            sh_ledger['R5'] = 'Portfolio Value'
            sh_ledger['S5'] = sh_portfolio['A4'].value

            self.risk_table_total_investment = sh_ledger['L11'].value

            sh_ledger['R4'] = 'Ideal Minimum Value'
            sh_ledger['S4'] = self.risk_table_total_investment

            # Write Headings and Rows
            headings = ['Risk', 'Allocation %', 'Allocation Value', 'Current %', 'Current Value', 'To reduce/add']
            # Write headings starting at R6 (Col 18)
            for col_idx, val in enumerate(headings, start=18):
                sh_ledger.cell(row=6, column=col_idx).value = val

            # Write Low Risk / High Risk rows
            sh_ledger['R7'], sh_ledger['S7'] = 'Low Risk', 0.5
            sh_ledger['R8'], sh_ledger['S8'] = 'High Risk', 0.5

            # Formatting (Percentage)
            pct_format = '0.00%'
            sh_ledger['S7'].number_format = pct_format
            sh_ledger['S8'].number_format = pct_format

            # Formulae
            
            for r in range(5, 10):
                sh_ledger[f'O{r}'] = f'=M{r}*S$5'

            # Risk Calculations
            self.portfolio_df['Risk'] = self.portfolio_df['Risk'].astype(str)
            
            low_risk_stocks = self.portfolio_df[self.portfolio_df['Risk'].str.upper() == "LOW"]
            high_risk_stocks = self.portfolio_df[self.portfolio_df['Risk'].str.upper() == "HIGH"]
            
            low_risk_stocks_sum = float(sh_funds['A16'].value or 0)
            high_risk_stocks_sum = float(sh_funds['A19'].value or 0)

            low_risk_stocks_sum += low_risk_stocks['Investment Value'].sum()
            high_risk_stocks_sum += high_risk_stocks['Investment Value'].sum()

            sh_ledger['V7'] = low_risk_stocks_sum
            sh_ledger['V8'] = high_risk_stocks_sum

            logger.debug(low_risk_stocks_sum)

            sh_ledger['U7'] = low_risk_stocks_sum/sh_portfolio['A4'].value #'=V7/S5'
            sh_ledger['U8'] = high_risk_stocks_sum/sh_portfolio['A4'].value #'=V8/S5'
            sh_ledger['U7'].number_format = pct_format
            sh_ledger['U8'].number_format = pct_format

            sh_ledger['T7'] = 0.5 * self.risk_table_total_investment #'=S7*S4'
            sh_ledger['T8'] = 0.5 * self.risk_table_total_investment #'=S8*S4'
            sh_ledger['W7'] = 0.5 * self.risk_table_total_investment - low_risk_stocks_sum #'=T7-V7'
            sh_ledger['W8'] = 0.5 * self.risk_table_total_investment - high_risk_stocks_sum #'=T8-V8'

            # Autofit
            for col in ['R', 'S', 'T', 'U', 'V', 'W']:
                sh_ledger.column_dimensions[col].auto_size = True 

            self.safe_save(wb)

            logger.info("Risk table updated.")
        
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)


    def update_beta_sheet(self, wb=None):
        logger.info("Updating Beta Sheet...")

        if wb is None:
            wb = self.safe_load()
        
        try:
            sh_portfolio = wb['Portfolio']
            sh_beta = wb['Beta']
            
            # Get Portfolio Weights
            last_row = sh_portfolio.max_row
            while last_row > 1 and sh_portfolio.cell(row=last_row, column=3).value is None:
                last_row -= 1
            
            # Reading Portfolio to get Allocation column
            
            portfolio_weight = self.portfolio_df['Allocation'].tolist()

            # Update Beta Sheet headers and weights
            for idx, (name, weight) in enumerate(zip(self.stock_names, portfolio_weight), start=11): # Col K is 11
                sh_beta.cell(row=3, column=idx).value = name
                sh_beta.cell(row=5, column=idx).value = weight
        
        except Exception as e:
            logger.debug("Error", exc_info=True)
        
        try:
            last_update = sh_beta['J8'].value
            stocks_history_download = yf.download('^NSEI', start=last_update)
            stocks_history_download = stocks_history_download['Close']
            stocks_history_download = stocks_history_download.reset_index()

            prices = stocks_history_download['^NSEI'].tolist()
            dates = stocks_history_download['Date'].tolist()
        
        except Exception as e:
            logger.debug("Error", exc_info=True)      
        
        try:
            # Write Dates and Prices (Vertical Transpose)
            for i, date_val in enumerate(dates):
                
                sh_beta.cell(row=8 + i, column=6).value = date_val.strftime('%Y-%m-%d')
                
                sh_beta.cell(row=8 + i, column=7).value = prices[i]

            sh_beta['J5'], sh_beta['F7'], sh_beta['G7'], sh_beta['I7'], sh_beta['H7'] = \
                'Portfolio Weight', 'Date', 'NIFTY 50', 'Portfolio Returns', 'NIFTY Returns'

            # Write Formulas
            for r in range(9, 8 + len(dates)):
                sh_beta[f'H{r}'] = f"=G{r}/G{r-1} - 1"
                sh_beta[f'H{r}'].number_format = '0.00%'
            
            self.update_sheets(wb)
        
        except Exception as e:
            logger.debug("Error", exc_info=True)

        try:
            sh_stock = wb[self.stock_names[0]]
            s_last = sh_stock.max_row
            while s_last > 1 and wb[self.stock_names[0]].cell(row=s_last, column=8).value is None:
                s_last -= 1

            # Loop through individual Stock Sheets
            start_col = 11 # Column K

            logger.debug(self.beta_dict.keys())
            
            for name in self.stock_names:

                stock_data = self.beta_dict[name]
                
                dates = [row[0] for row in stock_data]

                prices = [row[1] for row in stock_data]
                changes = [row[2] if len(row) > 2 else 0 for row in stock_data]

                temp_df = pd.DataFrame({"Date":dates, "Closing Price":prices, "Change %": changes})

                temp_df['Date'] = pd.to_datetime(temp_df['Date'], errors='coerce')

                temp_df["Change %"] = pd.to_numeric(temp_df["Change %"], errors='coerce')
                
                # Merge with NIFTY dates
                temp_df_merged = stocks_history_download.merge(temp_df, on='Date', how='left')
                
                logger.debug(temp_df.tail())
                sh_beta.cell(row=7, column=start_col).value = name
                
                # Write 'Change %' vertically
                changes = temp_df_merged['Change %'].tolist()
                for i, val in enumerate(changes):
                    sh_beta.cell(row=8 + i, column=start_col).value = val
                    sh_beta.cell(row=8 + i, column=start_col).number_format = "0.000%"
                
                start_col += 1
            self.safe_save(wb)

            # Final Headers and Datesx
            sh_beta['J6'], sh_beta['J7'] = 'Beta', 'Date'
            for i, date_val in enumerate(temp_df_merged['Date']):
                sh_beta.cell(row=8 + i, column=10).value = date_val.strftime('%Y-%m-%d')
            
            self.safe_save(wb)

            logger.info("Done updating Beta sheet!")

        except Exception as e:
            print("Caught an error")
            logger.debug("Error", exc_info=True)
        
    
    def update_beta_sheet_wrapper(self, wb=None):
        try:
            if wb is None:
                wb = self.safe_load()
            
            self.is_updating = True

            self.update_beta_sheet(wb)
        
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)
        finally:
            self.is_updating = False
        
            
        

    def update_ledger(self, time_period, investor_dict, wb=None):
        try:
            if wb is None:
                wb = self.safe_load()

            self.is_updating = True

            sh_ledger = wb['Ledger']

            last_row = sh_ledger.max_row
            while last_row > 1 and sh_ledger.cell(row=last_row, column=1).value is None:
                last_row -= 1
            
            i = 0
            for investor in self.investor_df_og['Investor']:
                money_added = investor_dict[investor]
                
                target_cell = sh_ledger.cell(row=last_row + 1, column=i + 2)
                if money_added > 0:
                    target_cell.value = "Paid"
                else:
                    target_cell.value = "Unpaid"
                    
                self.money_added_list[i] = self.money_added_list[i] + money_added
                self.ledger_dict[investor] = self.ledger_dict[investor] + money_added
                i += 1

            investor_total_investment = sum(self.money_added_list)

            self.investor_df_og['Amount Invested'] = self.money_added_list
            self.investor_df_og['% Total Fund'] = self.investor_df_og['Amount Invested']/investor_total_investment
            self.investor_df_og['Promised Amount'] = self.investor_df_og['Amount Invested'] * 1.14
            self.investor_df_og['Investment Value'] = self.investor_df_og['% Total Fund'] * self.portfolio_value
            self.investor_df_og['Investment Value'] = self.investor_df_og['Investment Value'].fillna(0).astype(float)
            self.investor_df_og['Profit/Loss'] = self.investor_df_og['Investment Value']/self.investor_df_og['Amount Invested'] -1

            investor_table_df = self.investor_df_og[['Amount Invested','% Total Fund',
            'Promised Amount', 'Investment Value', 'Profit/Loss']]

            for i, row in enumerate(investor_table_df.itertuples(), start=1):
                sh_ledger.cell(row=4 + i, column=12).value = row[1]
                sh_ledger.cell(row=4 + i, column=13).value = row[2]
                sh_ledger.cell(row=4 + i, column=13).number_format = "0.00%"
                sh_ledger.cell(row=4 + i, column=14).value = row[3]
                sh_ledger.cell(row=4 + i, column=15).value = row[4]
                sh_ledger.cell(row=4 + i, column=16).value = row[5]
                sh_ledger.cell(row=4 + i, column=16).number_format = "0.00%"
            
            #Adding sum of money invested to the investor's table
            sh_ledger.cell(row=11, column=12).value = investor_total_investment
            sh_ledger.cell(row=11, column=14).value = self.investor_df_og['Promised Amount'].sum()
            sh_ledger.cell(row=11, column=15).value = self.investor_df_og['Investment Value'].sum()

            #Updating "Ideal Minimum Value"
            sh_ledger.cell(row=4, column=19).value = investor_total_investment

            # Update A{last_row + 1} with time_period
            sh_ledger.cell(row=last_row + 1, column=1).value = time_period

            self.safe_save(wb)

            return f"Updated Ledger for the period {time_period}"
        
        except Exception as e:
            logger.debug("Error", exc_info=True)
            logger.info("Caught an error in update_ledger.")
            return "Error with Ledger Update."
        
        finally:
            self.is_updating = False



    def update_ledger_mini(self, wb):
        try:
            sh_ledger = wb['Ledger']
            self.investor_df_og['Investment Value'] = self.investor_df_og['% Total Fund'] * self.portfolio_value
            self.investor_df_og['Investment Value'] = self.investor_df_og['Investment Value'].fillna(0).astype(float)
            self.investor_df_og['Profit/Loss'] = self.investor_df_og['Investment Value']/self.investor_df_og['Amount Invested'] -1
            
            investor_table_df = self.investor_df_og[['Amount Invested','% Total Fund',
                'Promised Amount', 'Investment Value', 'Profit/Loss']]

            for i, row in enumerate(investor_table_df.itertuples(), start=1):
                sh_ledger.cell(row=4 + i, column=15).value = row[4]
                sh_ledger.cell(row=4 + i, column=16).value = row[5]
                sh_ledger.cell(row=4 + i, column=16).number_format = "0.00%"

            sh_ledger.cell(row=11, column=15).value = self.portfolio_value

            investor_total_investment = sum(self.money_added_list)
             
            sh_ledger.cell(row=4, column=19).value = investor_total_investment
            sh_ledger.cell(row=11, column=16).value = '=O11/L11 -1'
            sh_ledger.cell(row=11, column=16).number_format = "0.00%"

            self.safe_save(wb)

        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)



    def update_log(self, wb=None):
        logger.info("Starting Log Update...")
        try:
            if wb is None:
                wb = self.safe_load()
            
            self.is_updating = True

            sh_log = wb['Log']
            sh_port = wb['Portfolio']

            # Find the last row in Column C
            last_row = sh_log.max_row
            while last_row > 1 and sh_log.cell(row=last_row, column=3).value is None:
                last_row -= 1
            
            # Store previous values for calculations
            prev_date = sh_log.cell(row=last_row, column=3).value
            prev_portfolio_value = float(sh_log.cell(row=last_row, column=4).value or 0)
            prev_nifty_price = float(sh_log.cell(row=last_row, column=5).value or 0) # Avoid div by zero
            prev_nav_val = float(sh_log.cell(row=last_row, column=6).value or 0)

            # Get NAV 
            nav_val = float(self.portfolio_value/sh_port['A10'].value)

            # Prepare new data
            new_row = last_row + 1
            date_and_time = datetime.datetime.now(ZoneInfo(timezone))

            # Get Nifty Price 
            nifty = yf.Ticker('^NSEI')
            nifty_price = nifty.fast_info['last_price']

            # Calculations
            port_change = (self.portfolio_value / prev_portfolio_value - 1) if prev_portfolio_value else 0
            nifty_change = (nifty_price / prev_nifty_price - 1) if prev_nifty_price else 0
            nav_change = (nav_val / prev_nav_val - 1) if prev_nav_val else 0

            # Write to the New Row
            sh_log.cell(row=new_row, column=3).value = date_and_time.replace(tzinfo=None)
            sh_log.cell(row=new_row, column=4).value = float(self.portfolio_value)
            sh_log.cell(row=new_row, column=5).value = float(nifty_price)
            sh_log.cell(row=new_row, column=6).value = float(nav_val)
            sh_log.cell(row=new_row, column=7).value = port_change
            sh_log.cell(row=new_row, column=8).value = nifty_change
            sh_log.cell(row=new_row, column=9).value = nav_change

            # Apply percentage formatting to the change columns
            for col in range(7, 10):
               sh_log.cell(row=new_row, column=col).number_format = '0.00%'

            if wb is not None:
                self.safe_save(wb)
            
            new_row_data = [date_and_time.replace(tzinfo=None), float(self.portfolio_value), float(nifty_price),
                            float(nav_val), port_change, nifty_change, nav_change]

            self.log_df.loc[len(self.log_df)] = new_row_data

            logger.info("Log Updated")

            return f"The log was last updated on: {prev_date}. Log has been updated!"

        except Exception as e:
            logger.debug("Error", exc_info=True)
            return "Error updating log."
        
        finally:
            self.is_updating = False



    def update_portfolio(self, wb=None):
        try:
            logger.info("Starting Portfolio Update")

            if wb is None:
                wb = self.safe_load()
            
            self.is_updating = True

            sh_port = wb['Portfolio']
            sh_funds = wb['Funds_Portfolio']

            self.money_invested_sum = float(sh_funds['A4'].value)
            stock_tickers = ', '.join(self.stock_names)
            sh_port['A12'] = "Returns"
            sh_port['A13'] = '=A4/A7 -1'
            sh_port['A13'].number_format = "0.00%"
            
            self.add_data_to_portfolio(wb)

            self.risk_table(wb)

            self.update_ledger_mini(wb)

            self.df_for_dash(wb)

            logger.info("Update completed :)")
        
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)
        
        finally:
            self.is_updating = False



    def update_portfolio_dashboard(self, wb=None):
        try:
            if wb is None:
                wb = self.safe_load()

            sh_funds = wb['Funds_Portfolio']
            sh_port = wb['Portfolio']

            current_price, investment_value, total_profit_loss = [[] for i in range(3)]
            self.portfolio_value = float(sh_funds['A7'].value)
            total_profit = float(sh_funds['A10'].value)

            i = 0
            for name in self.stock_names:
                temp_stock = yf.Ticker(name)
                temp_price = temp_stock.get_info()['regularMarketPrice']
                current_price.append(temp_price)
                investment_value.append(temp_price * self.quantity_list[i])
                total_profit_loss.append(investment_value[-1] - self.money_invested[i])
                i += 1

            self.portfolio_value += sum(investment_value)
            total_profit += sum(total_profit_loss)

            sh_port['A14'] = total_profit
            sh_port['A4'] = self.portfolio_value

            todays_returns = float(self.portfolio_value/sh_port['A7'].value - 1)
            nav = float(self.portfolio_value/sh_port['A10'].value)

            return self.portfolio_value, todays_returns, nav
        
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)



    def update_sheets(self, wb=None):
        try:
            logger.info('Updating Individual Stock Sheets...')
            
            if wb is None:
                wb = self.safe_load()

            today_str = datetime.datetime.now(ZoneInfo(timezone)).strftime("%Y-%m-%d")
        
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)

        try:
            for name in self.stock_names:
                sh = wb[name]

                # Find the last row in Column H (Date column)
                last_row = sh.max_row
                while last_row > 1 and sh.cell(row=last_row, column=8).value is None:
                    last_row -= 1
                
                last_update = sh.cell(row=last_row, column=8).value

                # Check if we need to update
                if isinstance(last_update, datetime.datetime):
                    last_update_str = last_update.strftime("%Y-%m-%d")
                else:
                    last_update_str = str(last_update)
                

                if today_str != last_update_str:
                    
                    df = yf.download(name, start=last_update_str)
                    if df.empty:
                        continue
                    
                    stock_data = self.beta_dict[name]

                    df = df['Close']
                    df = df.reset_index()

                    df['Change %'] = df[name].shift(1) / df[name] - 1
                    
                    update_data = df[1:] if len(df) > 1 else pd.DataFrame()
                    
                    if update_data.empty:
                        continue
                    else:
                        update_data_list = list(update_data.itertuples(index=False, name=None))
                        stock_data.extend(update_data_list)
                        self.beta_dict[name] = stock_data  
                    
                    # Write Data (Vertical Transpose)
                    for i, row in enumerate(update_data.itertuples(), start=1):
                        new_row_idx = last_row + i
                        sh.cell(row=new_row_idx, column=8).value = row.Date
                        sh.cell(row=new_row_idx, column=9).value = row[2]
                        sh.cell(row=new_row_idx, column=9).number_format = "0.00"

                        # Apply Formulas and Formatting for the new rows
                        sh.cell(row=new_row_idx, column=10).value = f"=I{new_row_idx}/I{new_row_idx-1} - 1"
                        sh.cell(row=new_row_idx, column=10).number_format = "0.00%"
            
                self.safe_save(wb)  

            logger.info("Done updating sheets!")

        except Exception as e:
            logger.debug("Error", exc_info=True)

        

    def update_sheets_thread(self, wb=None):
        """Wrapper to handle the updates of sheets while ensuring safe writes"""
        try:
            self.is_updating = True
            if wb is None:
                wb = self.safe_load()
            self.update_sheets(wb) 
        finally:
            self.is_updating = False
        

    
    def update_transactions(self, date, symbol, action, quantity, price, sector_input=None, risk_input=None):
        try:
            wb = self.safe_load(data_only=False)
            sh_trans = wb['Transaction History']
            sh_port = wb['Portfolio']

            logger.info(f'Adding {action} entry for {symbol}...')

            quantity = float(quantity)
            if action == 'SELL':
                quantity = -1 * quantity
            
            amount = quantity * float(price)
            logger.debug(f"Amount {amount}")

            # Finding Right-most value in Transaction History
            right_most_col = 14
            while sh_trans.cell(row=5, column=right_most_col + 1).value is not None:
                right_most_col += 1
            
            self.money_invested_sum = float(sh_port['A7'].value or 0)
            
            self.money_invested_sum += amount
            
            # New Stock Sheet Creation
            if symbol not in self.stocks_dict:
                logger.info(f"New stock detected: {symbol}. Creating sheet...")
                
                # Update class attributes
                self.sector.append(sector_input)
                self.risk.append(risk_input)
                self.money_invested.append(amount)
                self.stocks_dict[symbol] = [quantity, amount]
                self.stock_names.append(symbol)
                self.quantity_list.append(quantity)
                self.start_dates.append(date)

                # Create new sheet
                new_sh = wb.create_sheet(title=symbol)
                new_sh['D3'], new_sh['E3'] = 'Name', symbol
                new_sh['D4'], new_sh['E4'] = 'First Purchased on', date
                new_sh['D5'], new_sh['E5'] = 'Risk', risk_input

                stocks_download = yf.download(symbol, start=self.portfolio_start_date)
                
                stocks_download = stocks_download['Close'].reset_index()

                # Write Headings
                headings = ['Date', 'Closing Price', 'Change %']
                for i, h in enumerate(headings, start=8): # H, I, J
                    new_sh.cell(row=3, column=i).value = h
                    
                # Write Dates and Prices (Vertical)
                for i, row in enumerate(stocks_download.itertuples(), start=4):
                    new_sh.cell(row=i, column=8).value = row.Date
                    new_sh.cell(row=i, column=9).value = getattr(row, symbol, row.Close)
                    if i > 4:
                        new_sh.cell(row=i, column=10).value = f"=I{i}/I{i-1} - 1"
                        new_sh.cell(row=i, column=10).number_format = "0.00%"

                # Add symbol to Transaction History headers
                sh_trans.cell(row=5, column=right_most_col + 1).value = symbol
                sh_trans.cell(row=6, column=right_most_col + 1).value = quantity

            else:
                # Update existing stock in dict
                self.stocks_dict[symbol][0] += quantity
                self.stocks_dict[symbol][1] += amount

            # Log the Transaction in 'Transaction History' 
            last_row_trans = sh_trans.max_row
            while last_row_trans > 1 and sh_trans.cell(row=last_row_trans, column=5).value is None:
                last_row_trans -= 1
            
            new_row = last_row_trans + 1
            sh_trans.cell(row=new_row, column=5).value = date
            sh_trans.cell(row=new_row, column=6).value = symbol
            sh_trans.cell(row=new_row, column=7).value = action
            sh_trans.cell(row=new_row, column=8).value = quantity
            sh_trans.cell(row=new_row, column=9).value = price
            sh_trans.cell(row=new_row, column=10).value = amount

            # Update Portfolio Sheet 
            sh_port['A7'] = self.money_invested_sum

            #Check for Sold-out Stocks (Quantity == 0)
            to_remove = [s for s in self.stock_names if self.stocks_dict[s][0] <= 0]
            
            if to_remove:
                for stock in to_remove:
                    self.stocks_dict.pop(stock, None)
                    logger.info(f"Removing {stock} from active portfolio.")
                
                # Filter the DataFrame
                self.portfolio_df = self.portfolio_df[~self.portfolio_df['Symbol'].isin(to_remove)]
                
                # Clearing Portfolio Range (C3:S{max}) before re-writing 
                for r in sh_port.iter_rows(min_row=3, max_row=sh_port.max_row, min_col=3, max_col=19):
                    for cell in r:
                        cell.value = None
            
                for i, name in enumerate(self.stock_names):
                    sh_trans.cell(row=5, column=14 + i).value = None
                    sh_trans.cell(row=6, column=14 + i).value = None
                
                self.start_dates = self.portfolio_df["Start Date"].to_list()
                self.money_invested = self.portfolio_df["Money Invested"].to_list()
                self.quantity_list = self.portfolio_df["Amount"].to_list()
                self.sector = self.portfolio_df["Sector"].to_list()
                self.risk = self.portfolio_df["Risk"].to_list()

            self.stock_names = self.portfolio_df["Symbol"].to_list()

            # Final Headers in Transaction History
            for i, name in enumerate(self.stock_names):
                sh_trans.cell(row=5, column=14 + i).value = name
                sh_trans.cell(row=6, column=14 + i).value = self.stocks_dict[name][0]

            # Update Portfolio and Risk 
            self.safe_save(wb)
            self.update_portfolio(wb)
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)
        
    
    def update_transactions_wrapper(self, date, symbol, action, quantity, price, sector_input=None, risk_input=None):
        try:
            self.is_updating = True
            self.update_transactions(date, symbol, action, quantity, price, sector_input=None, risk_input=None)
        except Exception as e:
            logger.debug(f"Error {e}", exc_info=True)
        finally:
            self.is_updating = False
